from openpyxl import load_workbook

wb = load_workbook("df_tabela.xlsx")
ws = wb["Sheet1"]
print(len(ws["A2"].value))



